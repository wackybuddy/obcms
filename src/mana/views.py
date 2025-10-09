from datetime import datetime, timedelta

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db import transaction
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from common.utils.moa_permissions import moa_no_access
from communities.models import OBCCommunity

from .models import (
    Assessment,
    AssessmentCategory,
    AssessmentTeamMember,
    MANAReport,
    Need,
    NeedsCategory,
    Survey,
    WorkshopActivity,
    WorkshopOutput,
    WorkshopParticipant,
    WorkshopSession,
)


@moa_no_access
@login_required
def new_assessment(request):
    """Create a new OBC-MANA assessment."""
    if request.method == "POST":
        try:
            with transaction.atomic():
                # Get form data
                title = request.POST.get("title")
                community_id = request.POST.get("community")
                primary_methodology = request.POST.get("primary_methodology")
                priority = request.POST.get("priority", "medium")
                assessment_level = request.POST.get("assessment_level", "community")
                planned_start_date = request.POST.get("planned_start_date")
                planned_end_date = request.POST.get("planned_end_date")
                objectives = request.POST.get("objectives")
                description = request.POST.get("description")
                estimated_budget = request.POST.get("estimated_budget")

                # Validate required fields
                if not all(
                    [
                        title,
                        community_id,
                        primary_methodology,
                        planned_start_date,
                        planned_end_date,
                        objectives,
                        description,
                    ]
                ):
                    messages.error(request, "Please fill in all required fields.")
                    return redirect("common:mana_new_assessment")

                # Get community
                community = get_object_or_404(OBCCommunity, id=community_id)

                # Get or create assessment category
                category, created = AssessmentCategory.objects.get_or_create(
                    name="OBC-MANA Workshop",
                    defaults={
                        "category_type": "needs_assessment",
                        "description": "Other Bangsamoro Communities Mapping and Needs Assessment",
                        "icon": "fas fa-users",
                        "color": "#3B82F6",
                    },
                )

                # Create assessment
                assessment = Assessment.objects.create(
                    title=title,
                    category=category,
                    description=description,
                    objectives=objectives,
                    assessment_level=assessment_level,
                    primary_methodology=primary_methodology,
                    community=community,
                    priority=priority,
                    planned_start_date=planned_start_date,
                    planned_end_date=planned_end_date,
                    estimated_budget=estimated_budget if estimated_budget else None,
                    lead_assessor=request.user,
                    created_by=request.user,
                )

                # Add team members
                team_leader_id = request.POST.get("team_leader")
                if team_leader_id:
                    AssessmentTeamMember.objects.create(
                        assessment=assessment,
                        user_id=team_leader_id,
                        role="team_leader",
                    )

                deputy_leader_id = request.POST.get("deputy_leader")
                if deputy_leader_id:
                    AssessmentTeamMember.objects.create(
                        assessment=assessment,
                        user_id=deputy_leader_id,
                        role="deputy_leader",
                    )

                primary_facilitator_id = request.POST.get("primary_facilitator")
                if primary_facilitator_id:
                    AssessmentTeamMember.objects.create(
                        assessment=assessment,
                        user_id=primary_facilitator_id,
                        role="facilitator",
                    )

                lead_documenter_id = request.POST.get("lead_documenter")
                if lead_documenter_id:
                    AssessmentTeamMember.objects.create(
                        assessment=assessment,
                        user_id=lead_documenter_id,
                        role="documenter",
                    )

                # If workshop methodology, create workshop activities
                if primary_methodology == "workshop":
                    create_workshop_activities(assessment, request.POST)

                messages.success(request, f'Assessment "{title}" created successfully!')
                return redirect("common:mana_manage_assessments")

        except Exception as e:
            messages.error(request, f"Error creating assessment: {str(e)}")
            return redirect("common:mana_new_assessment")

    # GET request - show form
    context = {
        "communities": OBCCommunity.objects.select_related(
            "barangay__municipality__province"
        ).all(),
        "users": User.objects.filter(is_active=True).order_by(
            "first_name", "last_name"
        ),
        "recent_assessments": Assessment.objects.select_related(
            "community__barangay"
        ).order_by("-created_at")[:5],
    }
    return render(request, "common/mana_new_assessment.html", context)


def create_workshop_activities(assessment, post_data):
    """Create the 5 core workshop activities for a workshop-based assessment."""

    start_date = datetime.strptime(
        post_data.get("planned_start_date"), "%Y-%m-%d"
    ).date()

    # Workshop definitions based on OBC-MANA guidelines
    workshops = [
        {
            "type": "workshop_1",
            "title": "Understanding the Community Context",
            "day": "day_2",
            "description": "Establish comprehensive understanding of OBC history, demographics, and socioeconomic conditions. Identify key strengths, resources, and assets within communities.",
            "methodology": "Small group discussions, participatory mapping, historical timeline development, stakeholder identification",
            "expected_outputs": "Community maps, historical timelines, inventory of community strengths, stakeholder maps",
            "duration": 4.0,
            "start_time": "09:00",
            "end_time": "13:00",
        },
        {
            "type": "workshop_2",
            "title": "Community Aspirations and Priorities",
            "day": "day_3",
            "description": "Document community vision for development. Identify key needs across social, economic, cultural, and rights dimensions. Prioritize sectors requiring immediate intervention.",
            "methodology": "Vision board creation, thematic small group discussions, structured prioritization exercises, plenary presentations",
            "expected_outputs": "Community vision statements, comprehensive needs assessment, prioritized list of development concerns",
            "duration": 4.0,
            "start_time": "09:00",
            "end_time": "13:00",
        },
        {
            "type": "workshop_3",
            "title": "Community Collaboration and Empowerment",
            "day": "day_3",
            "description": "Assess existing community organizations and leadership structures. Identify opportunities for strengthening collaboration and develop strategies for community empowerment.",
            "methodology": "Organizational mapping, collaboration assessment matrices, group work on empowerment strategies, action planning",
            "expected_outputs": "Map of community organizations, analysis of collaboration gaps, strategies for community capacity building",
            "duration": 4.0,
            "start_time": "14:00",
            "end_time": "18:00",
        },
        {
            "type": "workshop_4",
            "title": "Community Feedback on Existing Initiatives",
            "day": "day_4",
            "description": "Gather feedback on current government programs and services. Document successful interventions and lessons learned. Identify implementation challenges and gaps.",
            "methodology": "Program inventory development, small group program analysis, structured evaluation, recommendation formulation",
            "expected_outputs": "Inventory of programs with effectiveness assessment, documentation of best practices, analysis of implementation challenges",
            "duration": 3.0,
            "start_time": "09:00",
            "end_time": "12:00",
        },
        {
            "type": "workshop_5",
            "title": "OBCs Needs, Challenges, Factors, and Outcomes",
            "day": "day_4",
            "description": "Conduct in-depth analysis of priority issues. Explore root causes and contributing factors. Assess differential impacts on community segments.",
            "methodology": "Problem tree analysis, impact assessment matrices, factor relationship mapping, scenario development",
            "expected_outputs": "Problem trees for priority issues, analysis of impacts on different groups, relationship maps between factors",
            "duration": 3.0,
            "start_time": "13:00",
            "end_time": "16:00",
        },
    ]

    target_participants = int(post_data.get("target_participants", 30))

    for i, workshop_def in enumerate(workshops):
        # Calculate the date for each workshop
        if workshop_def["day"] == "day_2":
            workshop_date = start_date + timedelta(days=1)
        elif workshop_def["day"] == "day_3":
            workshop_date = start_date + timedelta(days=2)
        elif workshop_def["day"] == "day_4":
            workshop_date = start_date + timedelta(days=3)
        else:
            workshop_date = start_date

        WorkshopActivity.objects.create(
            assessment=assessment,
            workshop_type=workshop_def["type"],
            title=workshop_def["title"],
            description=workshop_def["description"],
            workshop_day=workshop_def["day"],
            scheduled_date=workshop_date,
            start_time=workshop_def["start_time"],
            end_time=workshop_def["end_time"],
            duration_hours=workshop_def["duration"],
            target_participants=target_participants,
            methodology=workshop_def["methodology"],
            expected_outputs=workshop_def["expected_outputs"],
            created_by=assessment.created_by,
        )


@moa_no_access
@login_required
def assessment_detail(request, assessment_id):
    """View detailed information about an assessment."""
    assessment = get_object_or_404(Assessment, id=assessment_id)

    context = {
        "assessment": assessment,
        "workshops": assessment.workshop_activities.all().order_by(
            "workshop_day", "start_time"
        ),
        "team_members": assessment.team_members.through.objects.filter(
            assessment=assessment
        ).select_related("user"),
        "can_edit": request.user == assessment.created_by
        or request.user == assessment.lead_assessor,
    }

    return render(request, "mana/assessment_detail.html", context)


@moa_no_access
@login_required
def workshop_detail(request, workshop_id):
    """View detailed information about a workshop."""
    workshop = get_object_or_404(WorkshopActivity, id=workshop_id)

    context = {
        "workshop": workshop,
        "sessions": workshop.sessions.all().order_by("session_order"),
        "participants": workshop.participants.all().order_by(
            "participant_type", "name"
        ),
        "outputs": workshop.outputs.all().order_by("output_type"),
        "can_edit": request.user == workshop.assessment.created_by
        or request.user in workshop.facilitators.all(),
    }

    return render(request, "mana/workshop_detail.html", context)


@moa_no_access
@login_required
def add_workshop_participant(request, workshop_id):
    """Add a participant to a workshop."""
    workshop = get_object_or_404(WorkshopActivity, id=workshop_id)

    if request.method == "POST":
        try:
            WorkshopParticipant.objects.create(
                workshop=workshop,
                name=request.POST.get("name"),
                participant_type=request.POST.get("participant_type"),
                gender=request.POST.get("gender"),
                age_group=request.POST.get("age_group"),
                contact_info=request.POST.get("contact_info", ""),
                organization=request.POST.get("organization", ""),
                attendance_status=request.POST.get("attendance_status", "attended"),
                participation_notes=request.POST.get("participation_notes", ""),
            )

            # Update actual participants count
            workshop.actual_participants = workshop.participants.count()
            workshop.save()

            messages.success(request, "Participant added successfully!")
        except Exception as e:
            messages.error(request, f"Error adding participant: {str(e)}")

    return redirect("mana:workshop_detail", workshop_id=workshop_id)


@moa_no_access
@login_required
def add_workshop_output(request, workshop_id):
    """Add an output to a workshop."""
    workshop = get_object_or_404(WorkshopActivity, id=workshop_id)

    if request.method == "POST":
        try:
            WorkshopOutput.objects.create(
                workshop=workshop,
                output_type=request.POST.get("output_type"),
                title=request.POST.get("title"),
                description=request.POST.get("description"),
                content=request.POST.get("content"),
                file_path=request.POST.get("file_path", ""),
                created_by=request.user,
            )

            messages.success(request, "Workshop output added successfully!")
        except Exception as e:
            messages.error(request, f"Error adding output: {str(e)}")

    return redirect("mana:workshop_detail", workshop_id=workshop_id)


@moa_no_access
@login_required
def generate_mana_report(request, assessment_id):
    """Generate or view the MANA report for an assessment."""
    assessment = get_object_or_404(Assessment, id=assessment_id)

    # Get or create the MANA report
    report, created = MANAReport.objects.get_or_create(
        assessment=assessment,
        defaults={
            "title": f"OBC-MANA Report - {assessment.community.barangay.name} - {assessment.created_at.year}",
            "created_by": request.user,
        },
    )

    if request.method == "POST":
        # Update report content
        report.executive_summary = request.POST.get("executive_summary", "")
        report.context_background = request.POST.get("context_background", "")
        report.methodology = request.POST.get("methodology", "")
        report.social_development_findings = request.POST.get(
            "social_development_findings", ""
        )
        report.economic_development_findings = request.POST.get(
            "economic_development_findings", ""
        )
        report.cultural_development_findings = request.POST.get(
            "cultural_development_findings", ""
        )
        report.rights_protection_findings = request.POST.get(
            "rights_protection_findings", ""
        )
        report.priority_issues = request.POST.get("priority_issues", "")
        report.policy_recommendations = request.POST.get("policy_recommendations", "")
        report.program_development_opportunities = request.POST.get(
            "program_development_opportunities", ""
        )
        report.strategic_approaches = request.POST.get("strategic_approaches", "")
        report.stakeholder_roles = request.POST.get("stakeholder_roles", "")
        report.resource_requirements = request.POST.get("resource_requirements", "")
        report.monitoring_evaluation = request.POST.get("monitoring_evaluation", "")

        # Update status if provided
        new_status = request.POST.get("report_status")
        if new_status:
            report.report_status = new_status

        report.save()
        messages.success(request, "MANA report updated successfully!")
        return redirect("mana:generate_mana_report", assessment_id=assessment_id)

    context = {
        "assessment": assessment,
        "report": report,
        "workshops": assessment.workshop_activities.all().order_by("workshop_day"),
        "can_edit": request.user == assessment.created_by
        or request.user == assessment.lead_assessor,
    }

    return render(request, "mana/mana_report.html", context)


# ==================== PHASE 2: MANA INTEGRATION VIEWS ====================


@moa_no_access
@login_required
def assessment_tasks_board(request, assessment_id):
    """Display kanban board for assessment tasks organized by phase."""
    assessment = get_object_or_404(Assessment, id=assessment_id)

    # Get tasks related to this assessment
    from common.work_item_model import WorkItem
    from django.contrib.contenttypes.models import ContentType

    # Get ContentType for Assessment model
    assessment_ct = ContentType.objects.get_for_model(Assessment)

    all_tasks = (
        WorkItem.objects.filter(
            work_type=WorkItem.WORK_TYPE_TASK,
            content_type=assessment_ct,
            object_id=assessment.id
        )
        .select_related("created_by")
        .prefetch_related("assignees", "teams")
    )

    # Define assessment phases with display metadata
    phases = {
        "planning": {
            "label": "Planning",
            "icon": "fa-clipboard-list",
            "color": "blue",
            "tasks": all_tasks.filter(task_data__assessment_phase="planning"),
        },
        "data_collection": {
            "label": "Data Collection",
            "icon": "fa-database",
            "color": "indigo",
            "tasks": all_tasks.filter(task_data__assessment_phase="data_collection"),
        },
        "analysis": {
            "label": "Analysis",
            "icon": "fa-chart-line",
            "color": "purple",
            "tasks": all_tasks.filter(task_data__assessment_phase="analysis"),
        },
        "report_writing": {
            "label": "Report Writing",
            "icon": "fa-file-alt",
            "color": "pink",
            "tasks": all_tasks.filter(task_data__assessment_phase="report_writing"),
        },
        "review": {
            "label": "Review",
            "icon": "fa-check-circle",
            "color": "emerald",
            "tasks": all_tasks.filter(task_data__assessment_phase="review"),
        },
    }

    # Calculate totals
    tasks_by_phase = {
        "total": all_tasks.count(),
        "completed": all_tasks.filter(status=WorkItem.STATUS_COMPLETED).count(),
    }

    context = {
        "assessment": assessment,
        "phases": phases,
        "tasks_by_phase": tasks_by_phase,
    }

    return render(request, "mana/assessment_tasks_board.html", context)


@moa_no_access
@login_required
def assessment_calendar(request, assessment_id):
    """Display calendar view for assessment milestones, tasks, and events."""
    assessment = get_object_or_404(Assessment, id=assessment_id)

    context = {
        "assessment": assessment,
    }

    return render(request, "mana/assessment_calendar.html", context)


@moa_no_access
@login_required
def assessment_calendar_feed(request, assessment_id):
    """JSON feed for FullCalendar showing milestones, tasks, and events."""
    assessment = get_object_or_404(Assessment, id=assessment_id)
    events = []

    # Milestones (from assessment phase completion dates)
    milestones = []

    if assessment.planning_completion_date:
        milestones.append(
            {
                "title": "\u2713 Planning Complete",
                "date": assessment.planning_completion_date,
                "type": "milestone",
            }
        )

    if assessment.data_collection_end_date:
        milestones.append(
            {
                "title": "\u2713 Data Collection Complete",
                "date": assessment.data_collection_end_date,
                "type": "milestone",
            }
        )

    if assessment.analysis_completion_date:
        milestones.append(
            {
                "title": "\u2713 Analysis Complete",
                "date": assessment.analysis_completion_date,
                "type": "milestone",
            }
        )

    if assessment.report_completion_date:
        milestones.append(
            {
                "title": "\u2713 Report Complete",
                "date": assessment.report_completion_date,
                "type": "milestone",
            }
        )

    if assessment.review_completion_date:
        milestones.append(
            {
                "title": "\u2713 Review Complete",
                "date": assessment.review_completion_date,
                "type": "milestone",
            }
        )

    for milestone in milestones:
        events.append(
            {
                "id": f'milestone-{milestone["title"].lower().replace(" ", "-")}',
                "title": milestone["title"],
                "start": milestone["date"].isoformat(),
                "allDay": True,
                "backgroundColor": "#3b82f6",
                "borderColor": "#3b82f6",
                "editable": False,
                "extendedProps": {"type": "milestone"},
            }
        )

    # Tasks (with due dates)
    from common.work_item_model import WorkItem
    from django.contrib.contenttypes.models import ContentType

    # Get ContentType for Assessment model
    assessment_ct = ContentType.objects.get_for_model(Assessment)

    tasks = WorkItem.objects.filter(
        work_type=WorkItem.WORK_TYPE_TASK,
        content_type=assessment_ct,
        object_id=assessment.id,
        due_date__isnull=False
    ).select_related("created_by")

    for task in tasks:
        events.append(
            {
                "id": str(task.id),
                "title": task.title,
                "start": task.due_date.isoformat(),
                "allDay": True,
                "backgroundColor": "#10b981",
                "borderColor": "#10b981",
                "editable": True,
                "extendedProps": {
                    "type": "task",
                    "status": task.status,
                    "priority": task.priority,
                },
            }
        )

    # Events (coordination events related to assessment)
    from coordination.models import Event

    coordination_events = Event.objects.filter(related_assessment=assessment)

    for event in coordination_events:
        event_data = {
            "id": f"event-{event.id}",
            "title": event.title,
            "start": event.start_date.isoformat(),
            "allDay": event.is_all_day if hasattr(event, "is_all_day") else True,
            "backgroundColor": "#f97316",
            "borderColor": "#f97316",
            "editable": False,
            "extendedProps": {"type": "event"},
        }

        if event.end_date:
            event_data["end"] = event.end_date.isoformat()

        if (
            event.start_time and not event.is_all_day
            if hasattr(event, "is_all_day")
            else False
        ):
            event_data["start"] = (
                f"{event.start_date.isoformat()}T{event.start_time.isoformat()}"
            )

        events.append(event_data)

    return JsonResponse(events, safe=False)


@moa_no_access
@login_required
def needs_prioritization_board(request):
    """Interactive board for ranking and prioritizing community needs."""
    from common.models import Region, Province
    from coordination.models import PPA

    # Get all needs
    needs_queryset = (
        Need.objects.select_related(
            "category",
            "assessment",
            "community__barangay__municipality__province__region",
        )
        .prefetch_related("linked_ppa")
        .order_by("-priority_score", "-community_votes", "created_at")
    )

    # Apply filters
    filters = {}

    if request.GET.get("sector"):
        filters["sector"] = request.GET.get("sector")
        needs_queryset = needs_queryset.filter(category_id=filters["sector"])

    if request.GET.get("region"):
        filters["region"] = request.GET.get("region")
        needs_queryset = needs_queryset.filter(
            community__barangay__municipality__province__region_id=filters["region"]
        )

    if request.GET.get("urgency"):
        filters["urgency"] = request.GET.get("urgency")
        needs_queryset = needs_queryset.filter(urgency_level=filters["urgency"])

    if request.GET.get("funding") == "funded":
        filters["funding"] = "funded"
        needs_queryset = needs_queryset.filter(linked_ppa__isnull=False)
    elif request.GET.get("funding") == "unfunded":
        filters["funding"] = "unfunded"
        needs_queryset = needs_queryset.filter(linked_ppa__isnull=True)

    needs = list(needs_queryset)

    # Count funded vs unfunded
    funded_count = sum(1 for need in needs if need.linked_ppa.exists())
    unfunded_count = len(needs) - funded_count

    # Get filter options
    categories = NeedsCategory.objects.all()
    regions = Region.objects.all()

    context = {
        "needs": needs,
        "funded_count": funded_count,
        "unfunded_count": unfunded_count,
        "categories": categories,
        "regions": regions,
        "filters": filters,
    }

    return render(request, "mana/needs_prioritization_board.html", context)


@moa_no_access
@login_required
def needs_update_ranking(request):
    """Update ranking order for needs via AJAX."""
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "Invalid request method"}, status=400
        )

    try:
        import json

        data = json.loads(request.body)
        needs_data = data.get("needs", [])

        # Update priority_score based on new ranking
        for item in needs_data:
            need_id = item.get("id")
            new_rank = item.get("rank")

            if need_id and new_rank:
                need = Need.objects.filter(id=need_id).first()
                if need:
                    # Higher rank = higher priority score
                    need.priority_score = 1000 - new_rank
                    need.save(update_fields=["priority_score"])

        return JsonResponse({"success": True})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@moa_no_access
@login_required
def need_vote(request, need_id):
    """Register a community vote for a need via AJAX."""
    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "Invalid request method"}, status=400
        )

    try:
        from .models import NeedVote

        need = get_object_or_404(Need, id=need_id)

        # Check if user already voted
        existing_vote = NeedVote.objects.filter(
            need=need, voted_by=request.user
        ).first()

        if existing_vote:
            return JsonResponse(
                {"success": False, "error": "You have already voted for this need."},
                status=400,
            )

        # Create vote
        NeedVote.objects.create(
            need=need,
            voted_by=request.user,
            vote_type="support",
            voted_at=timezone.now(),
        )

        # Update community votes count
        need.community_votes = NeedVote.objects.filter(need=need).count()
        need.save(update_fields=["community_votes"])

        return JsonResponse({"success": True, "votes": need.community_votes})

    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)}, status=500)


@moa_no_access
@login_required
def needs_export(request):
    """Export needs to Excel format."""
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Font, PatternFill, Alignment

    # Apply same filters as prioritization board
    needs_queryset = (
        Need.objects.select_related(
            "category",
            "assessment",
            "community__barangay__municipality__province__region",
        )
        .prefetch_related("linked_ppa")
        .order_by("-priority_score", "-community_votes")
    )

    # Apply filters
    if request.GET.get("sector"):
        needs_queryset = needs_queryset.filter(category_id=request.GET.get("sector"))

    if request.GET.get("region"):
        needs_queryset = needs_queryset.filter(
            community__barangay__municipality__province__region_id=request.GET.get(
                "region"
            )
        )

    if request.GET.get("urgency"):
        needs_queryset = needs_queryset.filter(urgency_level=request.GET.get("urgency"))

    if request.GET.get("funding") == "funded":
        needs_queryset = needs_queryset.filter(linked_ppa__isnull=False)
    elif request.GET.get("funding") == "unfunded":
        needs_queryset = needs_queryset.filter(linked_ppa__isnull=True)

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Community Needs"

    # Header row
    headers = [
        "Rank",
        "Title",
        "Category",
        "Community",
        "Municipality",
        "Province",
        "Region",
        "Urgency",
        "Votes",
        "Budget Est. (PHP)",
        "Funding Status",
    ]

    header_fill = PatternFill(
        start_color="4F81BD", end_color="4F81BD", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for rank, need in enumerate(needs_queryset, start=1):
        ws.append(
            [
                rank,
                need.title,
                need.category.name if need.category else "",
                (
                    need.community.display_name
                    if hasattr(need.community, "display_name")
                    else need.community.barangay.name
                ),
                need.community.barangay.municipality.name,
                need.community.barangay.municipality.province.name,
                need.community.barangay.municipality.province.region.name,
                need.get_urgency_level_display(),
                need.community_votes or 0,
                need.estimated_cost if need.estimated_cost else "",
                "Funded" if need.linked_ppa.exists() else "Unfunded",
            ]
        )

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Create response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename=community_needs_{timezone.now().strftime("%Y%m%d")}.xlsx'
    )

    wb.save(response)
    return response
