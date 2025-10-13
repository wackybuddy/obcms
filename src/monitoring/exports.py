"""Export views for planning and budgeting reports."""

import csv
from datetime import datetime
from decimal import Decimal
from io import BytesIO

from django.contrib.auth.decorators import login_required
from django.db.models import Count, Q, Sum
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone

from common.decorators.rbac import require_feature_access
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import (
    MonitoringEntry,
    MonitoringEntryFunding,
    MonitoringEntryWorkflowStage,
)

ZERO_DECIMAL = Decimal("0.00")


@login_required
@require_feature_access('monitoring_access')
def export_aip_summary_excel(request):
    """Export Annual Investment Plan summary to Excel format."""
    # Get filter parameters
    plan_year = request.GET.get("plan_year")
    sector = request.GET.get("sector")
    funding_source = request.GET.get("funding_source")

    # Query entries
    entries = MonitoringEntry.objects.select_related(
        "lead_organization",
        "implementing_moa",
        "submitted_by_community",
        "coverage_region",
        "coverage_province",
    ).prefetch_related("funding_flows")

    if plan_year:
        entries = entries.filter(plan_year=plan_year)
    if sector:
        entries = entries.filter(sector=sector)
    if funding_source:
        entries = entries.filter(funding_source=funding_source)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "AIP Summary"

    # Header styling
    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Define headers
    headers = [
        "Category",
        "Title",
        "Lead Organization",
        "Implementing MOA",
        "Plan Year",
        "Fiscal Year",
        "Sector",
        "Appropriation Class",
        "Funding Source",
        "Program Code",
        "Budget Allocation (PHP)",
        "OBC Allocation (PHP)",
        "Budget Ceiling (PHP)",
        "Allocations (PHP)",
        "Obligations (PHP)",
        "Disbursements (PHP)",
        "Status",
        "Progress (%)",
        "Region",
        "Province",
        "GAD",
        "CCET",
        "IP",
        "Peace",
        "SDG",
        "Start Date",
        "Target End Date",
    ]

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data rows
    row_num = 2
    for entry in entries:
        # Calculate funding totals
        allocations = entry.funding_total(MonitoringEntryFunding.TRANCHE_ALLOCATION)
        obligations = entry.funding_total(MonitoringEntryFunding.TRANCHE_OBLIGATION)
        disbursements = entry.funding_total(MonitoringEntryFunding.TRANCHE_DISBURSEMENT)

        row_data = [
            entry.get_category_display(),
            entry.title,
            entry.lead_organization.name if entry.lead_organization else "",
            entry.implementing_moa.name if entry.implementing_moa else "",
            entry.plan_year or "",
            entry.fiscal_year or "",
            entry.get_sector_display() if entry.sector else "",
            (
                entry.get_appropriation_class_display()
                if entry.appropriation_class
                else ""
            ),
            entry.get_funding_source_display() if entry.funding_source else "",
            entry.program_code or "",
            float(entry.budget_allocation or 0),
            float(entry.budget_obc_allocation or 0),
            float(entry.budget_ceiling or 0),
            float(allocations),
            float(obligations),
            float(disbursements),
            entry.get_status_display(),
            entry.progress,
            entry.coverage_region.name if entry.coverage_region else "",
            entry.coverage_province.name if entry.coverage_province else "",
            "Yes" if entry.compliance_gad else "No",
            "Yes" if entry.compliance_ccet else "No",
            "Yes" if entry.benefits_indigenous_peoples else "No",
            "Yes" if entry.supports_peace_agenda else "No",
            "Yes" if entry.supports_sdg else "No",
            entry.start_date.strftime("%Y-%m-%d") if entry.start_date else "",
            entry.target_end_date.strftime("%Y-%m-%d") if entry.target_end_date else "",
        ]

        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

        row_num += 1

    # Add summary row
    summary_row = row_num + 1
    ws.cell(row=summary_row, column=1, value="TOTAL").font = Font(bold=True)

    # Sum budget columns (columns 11-16)
    for col in range(11, 17):
        col_letter = get_column_letter(col)
        formula = f"=SUM({col_letter}2:{col_letter}{row_num-1})"
        cell = ws.cell(row=summary_row, column=col, value=formula)
        cell.font = Font(bold=True)
        cell.number_format = "#,##0.00"

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Number formatting for currency columns
    for row in range(2, row_num):
        for col in range(11, 17):  # Budget columns
            ws.cell(row=row, column=col).number_format = "#,##0.00"

    # Generate response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"AIP_Summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_feature_access('monitoring_access')
def export_compliance_report_excel(request):
    """Export compliance tracking report (GAD, CCET, IP, Peace, SDG)."""
    entries = MonitoringEntry.objects.select_related(
        "lead_organization",
        "implementing_moa",
    ).all()

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Compliance Report"

    # Header styling
    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)

    headers = [
        "Title",
        "Category",
        "Lead Organization",
        "Sector",
        "Budget Allocation (PHP)",
        "GAD",
        "CCET",
        "Indigenous Peoples",
        "Peace Agenda",
        "SDG",
        "Goal Alignment",
        "Moral Governance Pillar",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Write data
    row_num = 2
    for entry in entries:
        row_data = [
            entry.title,
            entry.get_category_display(),
            entry.lead_organization.name if entry.lead_organization else "",
            entry.get_sector_display() if entry.sector else "",
            float(entry.budget_allocation or 0),
            "Yes" if entry.compliance_gad else "No",
            "Yes" if entry.compliance_ccet else "No",
            "Yes" if entry.benefits_indigenous_peoples else "No",
            "Yes" if entry.supports_peace_agenda else "No",
            "Yes" if entry.supports_sdg else "No",
            ", ".join(entry.goal_alignment) if entry.goal_alignment else "",
            entry.moral_governance_pillar or "",
        ]

        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=value)

        row_num += 1

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Format currency column
    for row in range(2, row_num):
        ws.cell(row=row, column=5).number_format = "#,##0.00"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Compliance_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_feature_access('monitoring_access')
def export_budget_csv(request):
    """Export budget summary as CSV."""
    entries = MonitoringEntry.objects.select_related(
        "lead_organization",
        "implementing_moa",
    ).all()

    response = HttpResponse(content_type="text/csv")
    filename = f"Budget_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow(
        [
            "Title",
            "Category",
            "Lead Organization",
            "Plan Year",
            "Fiscal Year",
            "Sector",
            "Appropriation Class",
            "Funding Source",
            "Budget Allocation",
            "OBC Allocation",
            "Budget Ceiling",
            "Status",
            "Progress",
            "GAD",
            "CCET",
            "IP",
            "Peace",
            "SDG",
        ]
    )

    for entry in entries:
        writer.writerow(
            [
                entry.title,
                entry.get_category_display(),
                entry.lead_organization.name if entry.lead_organization else "",
                entry.plan_year or "",
                entry.fiscal_year or "",
                entry.get_sector_display() if entry.sector else "",
                (
                    entry.get_appropriation_class_display()
                    if entry.appropriation_class
                    else ""
                ),
                entry.get_funding_source_display() if entry.funding_source else "",
                float(entry.budget_allocation or 0),
                float(entry.budget_obc_allocation or 0),
                float(entry.budget_ceiling or 0),
                entry.get_status_display(),
                entry.progress,
                "Yes" if entry.compliance_gad else "No",
                "Yes" if entry.compliance_ccet else "No",
                "Yes" if entry.benefits_indigenous_peoples else "No",
                "Yes" if entry.supports_peace_agenda else "No",
                "Yes" if entry.supports_sdg else "No",
            ]
        )

    return response


@login_required
@require_feature_access('monitoring_access')
def export_funding_timeline_excel(request):
    """Export funding timeline (allocations, obligations, disbursements)."""
    tranches = MonitoringEntryFunding.objects.select_related(
        "entry__lead_organization",
        "entry__implementing_moa",
    ).order_by("scheduled_date")

    wb = Workbook()
    ws = wb.active
    ws.title = "Funding Timeline"

    header_fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)

    headers = [
        "PPA Title",
        "Lead Organization",
        "Tranche Type",
        "Amount (PHP)",
        "Funding Source",
        "Scheduled Date",
        "Remarks",
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font

    row_num = 2
    for tranche in tranches:
        row_data = [
            tranche.entry.title,
            (
                tranche.entry.lead_organization.name
                if tranche.entry.lead_organization
                else ""
            ),
            tranche.get_tranche_type_display(),
            float(tranche.amount),
            tranche.get_funding_source_display() if tranche.funding_source else "",
            (
                tranche.scheduled_date.strftime("%Y-%m-%d")
                if tranche.scheduled_date
                else ""
            ),
            tranche.remarks or "",
        ]

        for col_num_inner, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num_inner, value=value)

        row_num += 1

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Format amount column
    for row in range(2, row_num):
        ws.cell(row=row, column=4).number_format = "#,##0.00"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Funding_Timeline_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
