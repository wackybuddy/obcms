"""
Government Compliance Reports Service.

This module provides report generation for:
1. MFBM (Ministry of Finance and Budget Management) - Budget Execution Report
2. BPDA (Bangsamoro Planning and Development Authority) - Development Alignment Report
3. COA (Commission on Audit) - Budget Variance and Audit Trail Report

All reports follow official government formats and include comprehensive audit trails.
"""

import io
from datetime import datetime
from decimal import Decimal
from typing import Optional

from auditlog.models import LogEntry
from django.contrib.contenttypes.models import ContentType
from django.db.models import Q, Sum
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from monitoring.models import MonitoringEntry, MonitoringEntryFunding


class MFBMBudgetExecutionReportGenerator:
    """
    MFBM Budget Execution Report Generator.

    Generates comprehensive budget execution reports following MFBM format standards,
    including PPA allocation, expenditure, variance analysis, and WorkItem tracking.
    """

    def __init__(self, fiscal_year: int, moa_filter: Optional[str] = None):
        """
        Initialize MFBM report generator.

        Args:
            fiscal_year: Fiscal year for the report
            moa_filter: Optional MOA organization ID to filter by specific ministry/office
        """
        self.fiscal_year = fiscal_year
        self.moa_filter = moa_filter

    def generate(self) -> io.BytesIO:
        """
        Generate MFBM Budget Execution Report in Excel format.

        Returns:
            BytesIO: Excel file buffer ready for download
        """
        # Query MOA PPAs for the fiscal year
        queryset = MonitoringEntry.objects.filter(
            category='moa_ppa',
            fiscal_year=self.fiscal_year
        ).select_related(
            'implementing_moa',
            'execution_project'
        ).prefetch_related(
            'funding_flows'
        )

        if self.moa_filter:
            queryset = queryset.filter(implementing_moa__id=self.moa_filter)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Budget Execution Report"

        # Write header section
        self._write_header(ws)

        # Write column headers
        row = 8
        headers = [
            "PPA Code",
            "PPA Title",
            "Implementing MOA",
            "Appropriation Class",
            "Allocated Budget (₱)",
            "Total Obligations (₱)",
            "Total Disbursements (₱)",
            "Variance (₱)",
            "Variance (%)",
            "Utilization Rate (%)",
            "Status",
            "WorkItem Expenditure (₱)",
            "Remarks"
        ]
        self._write_column_headers(ws, row, headers)

        # Write data rows
        row += 1
        total_allocated = Decimal('0.00')
        total_obligations = Decimal('0.00')
        total_disbursements = Decimal('0.00')
        total_workitem_expenditure = Decimal('0.00')

        for entry in queryset:
            allocated = entry.budget_allocation or Decimal('0.00')
            obligations = entry.total_obligations
            disbursements = entry.total_disbursements
            variance = allocated - disbursements
            variance_pct = (variance / allocated * 100) if allocated > 0 else 0
            utilization_rate = entry.budget_utilization_rate

            # Get WorkItem expenditure if tracking enabled
            workitem_expenditure = Decimal('0.00')
            if entry.enable_workitem_tracking and entry.execution_project:
                workitem_expenditure = entry.execution_project.actual_expenditure or Decimal('0.00')

            # Write row data
            ws.cell(row, 1, entry.program_code or f"PPA-{entry.id.hex[:8]}")
            ws.cell(row, 2, entry.title)
            ws.cell(row, 3, entry.implementing_moa.name if entry.implementing_moa else "")
            ws.cell(row, 4, entry.get_appropriation_class_display())
            ws.cell(row, 5, float(allocated))
            ws.cell(row, 6, float(obligations))
            ws.cell(row, 7, float(disbursements))
            ws.cell(row, 8, float(variance))
            ws.cell(row, 9, float(variance_pct))
            ws.cell(row, 10, float(utilization_rate))
            ws.cell(row, 11, entry.get_status_display())
            ws.cell(row, 12, float(workitem_expenditure))

            # Remarks based on variance
            remarks = self._generate_remarks(entry, variance, variance_pct)
            ws.cell(row, 13, remarks)

            # Apply number formatting
            for col in [5, 6, 7, 8, 12]:
                ws.cell(row, col).number_format = '#,##0.00'
            for col in [9, 10]:
                ws.cell(row, col).number_format = '0.00'

            # Accumulate totals
            total_allocated += allocated
            total_obligations += obligations
            total_disbursements += disbursements
            total_workitem_expenditure += workitem_expenditure

            row += 1

        # Write totals row
        self._write_totals_row(
            ws, row, total_allocated, total_obligations,
            total_disbursements, total_workitem_expenditure
        )

        # Auto-adjust column widths
        self._adjust_column_widths(ws)

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _write_header(self, ws):
        """Write report header section."""
        ws.merge_cells('A1:M1')
        ws['A1'] = "MINISTRY OF FINANCE AND BUDGET MANAGEMENT"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:M2')
        ws['A2'] = "BANGSAMORO AUTONOMOUS REGION IN MUSLIM MINDANAO"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A3:M3')
        ws['A3'] = f"BUDGET EXECUTION REPORT - FY {self.fiscal_year}"
        ws['A3'].font = Font(bold=True, size=12)
        ws['A3'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A5:M5')
        ws['A5'] = f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}"
        ws['A5'].font = Font(italic=True, size=10)
        ws['A5'].alignment = Alignment(horizontal='center')

    def _write_column_headers(self, ws, row, headers):
        """Write column headers with formatting."""
        header_fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

    def _write_totals_row(self, ws, row, allocated, obligations, disbursements, workitem_exp):
        """Write totals summary row."""
        ws.cell(row, 1, "TOTAL")
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 5, float(allocated))
        ws.cell(row, 6, float(obligations))
        ws.cell(row, 7, float(disbursements))
        ws.cell(row, 8, float(allocated - disbursements))
        ws.cell(row, 12, float(workitem_exp))

        # Calculate overall variance percentage
        overall_variance_pct = ((allocated - disbursements) / allocated * 100) if allocated > 0 else 0
        ws.cell(row, 9, float(overall_variance_pct))

        # Calculate overall utilization rate
        overall_utilization = (disbursements / allocated * 100) if allocated > 0 else 0
        ws.cell(row, 10, float(overall_utilization))

        # Apply number formatting
        for col in [5, 6, 7, 8, 12]:
            ws.cell(row, col).number_format = '#,##0.00'
            ws.cell(row, col).font = Font(bold=True)
        for col in [9, 10]:
            ws.cell(row, col).number_format = '0.00'
            ws.cell(row, col).font = Font(bold=True)

    def _generate_remarks(self, entry, variance, variance_pct):
        """Generate remarks based on variance analysis."""
        remarks = []

        if variance_pct > 20:
            remarks.append("High underspending")
        elif variance_pct < -5:
            remarks.append("Over budget")

        if entry.budget_utilization_rate < 50:
            remarks.append("Low utilization")
        elif entry.budget_utilization_rate > 95:
            remarks.append("Near full utilization")

        if entry.approval_status != 'approved':
            remarks.append(f"Status: {entry.get_approval_status_display()}")

        return "; ".join(remarks) if remarks else "On track"

    def _adjust_column_widths(self, ws):
        """Auto-adjust column widths for readability."""
        column_widths = {
            'A': 15, 'B': 40, 'C': 30, 'D': 20, 'E': 18,
            'F': 18, 'G': 18, 'H': 15, 'I': 12, 'J': 15,
            'K': 15, 'L': 20, 'M': 30
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width


class BPDADevelopmentReportGenerator:
    """
    BPDA Development Alignment Report Generator.

    Generates development alignment reports showing how PPAs align with
    the Bangsamoro Development Plan (BDP), including outcome indicators
    and beneficiary analysis.
    """

    def __init__(self, fiscal_year: int, sector: Optional[str] = None):
        """
        Initialize BPDA report generator.

        Args:
            fiscal_year: Fiscal year for the report
            sector: Optional sector filter (economic, social, infrastructure, etc.)
        """
        self.fiscal_year = fiscal_year
        self.sector = sector

    def generate(self) -> io.BytesIO:
        """
        Generate BPDA Development Alignment Report in Excel format.

        Returns:
            BytesIO: Excel file buffer ready for download
        """
        # Query PPAs for the fiscal year
        queryset = MonitoringEntry.objects.filter(
            fiscal_year=self.fiscal_year
        ).exclude(
            sector__isnull=True
        ).exclude(
            sector=''
        ).select_related(
            'implementing_moa',
            'lead_organization'
        ).prefetch_related(
            'standard_outcome_indicators',
            'communities'
        )

        if self.sector:
            queryset = queryset.filter(sector=self.sector)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Development Alignment"

        # Write header section
        self._write_header(ws)

        # Write column headers
        row = 8
        headers = [
            "PPA Code",
            "PPA Title",
            "Sector",
            "BDP Alignment Score",
            "Outcome Indicators",
            "Beneficiaries",
            "OBC Communities",
            "Status",
            "Progress (%)",
            "SDG Contribution",
            "Moral Governance Pillar",
            "Remarks"
        ]
        self._write_column_headers(ws, row, headers)

        # Write data rows
        row += 1
        for entry in queryset:
            # Calculate BDP alignment score
            alignment_score = self._calculate_bdp_alignment(entry)

            # Get outcome indicators
            outcome_indicators = ", ".join(
                [ind.indicator_name for ind in entry.standard_outcome_indicators.all()[:3]]
            )

            # Get beneficiary counts
            total_beneficiaries = entry.total_slots or 0
            obc_beneficiaries = entry.obc_slots or 0
            communities_count = entry.communities.count()

            # Write row data
            ws.cell(row, 1, entry.program_code or f"PPA-{entry.id.hex[:8]}")
            ws.cell(row, 2, entry.title)
            ws.cell(row, 3, entry.get_sector_display())
            ws.cell(row, 4, alignment_score)
            ws.cell(row, 5, outcome_indicators or "N/A")
            ws.cell(row, 6, f"{obc_beneficiaries}/{total_beneficiaries}")
            ws.cell(row, 7, communities_count)
            ws.cell(row, 8, entry.get_status_display())
            ws.cell(row, 9, entry.progress)
            ws.cell(row, 10, "Yes" if entry.supports_sdg else "No")
            ws.cell(row, 11, entry.moral_governance_pillar or "N/A")

            # Remarks based on alignment
            remarks = self._generate_development_remarks(entry, alignment_score)
            ws.cell(row, 12, remarks)

            # Apply number formatting
            ws.cell(row, 4).number_format = '0.00'
            ws.cell(row, 9).number_format = '0'

            row += 1

        # Add summary statistics
        self._write_summary_statistics(ws, row + 2, queryset)

        # Auto-adjust column widths
        self._adjust_column_widths(ws)

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _write_header(self, ws):
        """Write report header section."""
        ws.merge_cells('A1:L1')
        ws['A1'] = "BANGSAMORO PLANNING AND DEVELOPMENT AUTHORITY"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:L2')
        ws['A2'] = "BANGSAMORO AUTONOMOUS REGION IN MUSLIM MINDANAO"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A3:L3')
        ws['A3'] = f"DEVELOPMENT ALIGNMENT REPORT - FY {self.fiscal_year}"
        ws['A3'].font = Font(bold=True, size=12)
        ws['A3'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A5:L5')
        ws['A5'] = f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}"
        ws['A5'].font = Font(italic=True, size=10)
        ws['A5'].alignment = Alignment(horizontal='center')

    def _write_column_headers(self, ws, row, headers):
        """Write column headers with formatting."""
        header_fill = PatternFill(start_color='0F7C3E', end_color='0F7C3E', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

    def _calculate_bdp_alignment(self, entry) -> float:
        """
        Calculate BDP alignment score based on various factors.

        Score components:
        - Has outcome indicators: 25 points
        - Supports SDG: 15 points
        - Has moral governance pillar: 20 points
        - Benefits OBCs: 20 points
        - Has policy linkage: 20 points

        Returns:
            float: Alignment score from 0-100
        """
        score = 0.0

        # Outcome indicators (25 points)
        if entry.standard_outcome_indicators.exists():
            score += 25.0

        # SDG contribution (15 points)
        if entry.supports_sdg:
            score += 15.0

        # Moral governance alignment (20 points)
        if entry.moral_governance_pillar:
            score += 20.0

        # OBC beneficiary component (20 points)
        if entry.obc_slots and entry.total_slots:
            obc_ratio = entry.obc_slots / entry.total_slots
            score += obc_ratio * 20.0

        # Policy linkage (20 points)
        if entry.implementing_policies.exists():
            score += 20.0

        return round(score, 2)

    def _generate_development_remarks(self, entry, alignment_score) -> str:
        """Generate remarks based on development alignment."""
        remarks = []

        if alignment_score >= 80:
            remarks.append("Strong BDP alignment")
        elif alignment_score >= 60:
            remarks.append("Good alignment")
        elif alignment_score >= 40:
            remarks.append("Moderate alignment")
        else:
            remarks.append("Needs alignment review")

        if entry.benefits_indigenous_peoples:
            remarks.append("IP focus")

        if entry.supports_peace_agenda:
            remarks.append("Peace building")

        return "; ".join(remarks)

    def _write_summary_statistics(self, ws, start_row, queryset):
        """Write summary statistics section."""
        ws.cell(start_row, 1, "SUMMARY STATISTICS")
        ws.cell(start_row, 1).font = Font(bold=True, size=12)

        row = start_row + 2

        # Sector breakdown
        sector_stats = queryset.values('sector').annotate(count=Sum('id'))
        ws.cell(row, 1, "Sector Breakdown:")
        ws.cell(row, 1).font = Font(bold=True)
        row += 1

        for stat in sector_stats:
            sector_display = dict(MonitoringEntry.SECTOR_CHOICES).get(stat['sector'], stat['sector'])
            ws.cell(row, 2, f"{sector_display}: {stat['count']} PPAs")
            row += 1

        # SDG contribution
        row += 1
        sdg_count = queryset.filter(supports_sdg=True).count()
        ws.cell(row, 1, f"SDG Contributing PPAs: {sdg_count}")
        ws.cell(row, 1).font = Font(bold=True)

    def _adjust_column_widths(self, ws):
        """Auto-adjust column widths for readability."""
        column_widths = {
            'A': 15, 'B': 40, 'C': 20, 'D': 18, 'E': 35,
            'F': 15, 'G': 15, 'H': 15, 'I': 12, 'J': 15,
            'K': 25, 'L': 30
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width


class COAVarianceReportGenerator:
    """
    COA Budget Variance Report Generator.

    Generates comprehensive budget variance reports with activity-level breakdown
    and audit trail export for Commission on Audit compliance.
    """

    def __init__(self, fiscal_year: int, moa_filter: Optional[str] = None):
        """
        Initialize COA report generator.

        Args:
            fiscal_year: Fiscal year for the report
            moa_filter: Optional MOA organization ID filter
        """
        self.fiscal_year = fiscal_year
        self.moa_filter = moa_filter

    def generate(self) -> io.BytesIO:
        """
        Generate COA Budget Variance Report in Excel format.

        Returns:
            BytesIO: Excel file buffer with variance analysis and audit trail
        """
        # Query PPAs for the fiscal year
        queryset = MonitoringEntry.objects.filter(
            fiscal_year=self.fiscal_year
        ).select_related(
            'implementing_moa',
            'execution_project'
        ).prefetch_related(
            'funding_flows',
            'execution_project__children'
        )

        if self.moa_filter:
            queryset = queryset.filter(implementing_moa__id=self.moa_filter)

        # Create Excel workbook
        wb = Workbook()

        # Sheet 1: Budget Variance Summary
        ws_summary = wb.active
        ws_summary.title = "Budget Variance"
        self._write_variance_summary(ws_summary, queryset)

        # Sheet 2: Activity-Level Breakdown
        ws_activities = wb.create_sheet("Activity Breakdown")
        self._write_activity_breakdown(ws_activities, queryset)

        # Sheet 3: Audit Trail
        ws_audit = wb.create_sheet("Audit Trail")
        self._write_audit_trail(ws_audit, queryset)

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def _write_variance_summary(self, ws, queryset):
        """Write budget variance summary sheet."""
        # Header
        ws.merge_cells('A1:K1')
        ws['A1'] = "COMMISSION ON AUDIT"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:K2')
        ws['A2'] = f"BUDGET VARIANCE REPORT - FY {self.fiscal_year}"
        ws['A2'].font = Font(bold=True, size=12)
        ws['A2'].alignment = Alignment(horizontal='center')

        ws.merge_cells('A4:K4')
        ws['A4'] = f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}"
        ws['A4'].font = Font(italic=True, size=10)
        ws['A4'].alignment = Alignment(horizontal='center')

        # Column headers
        row = 7
        headers = [
            "PPA Code",
            "PPA Title",
            "Implementing MOA",
            "Budget Allocation (₱)",
            "Total Obligations (₱)",
            "Total Disbursements (₱)",
            "Variance (₱)",
            "Variance (%)",
            "Obligation Rate (%)",
            "Disbursement Rate (%)",
            "Audit Findings"
        ]
        self._write_column_headers(ws, row, headers)

        # Data rows
        row += 1
        for entry in queryset:
            allocated = entry.budget_allocation or Decimal('0.00')
            obligations = entry.total_obligations
            disbursements = entry.total_disbursements
            variance = allocated - disbursements
            variance_pct = (variance / allocated * 100) if allocated > 0 else 0

            ws.cell(row, 1, entry.program_code or f"PPA-{entry.id.hex[:8]}")
            ws.cell(row, 2, entry.title)
            ws.cell(row, 3, entry.implementing_moa.name if entry.implementing_moa else "")
            ws.cell(row, 4, float(allocated))
            ws.cell(row, 5, float(obligations))
            ws.cell(row, 6, float(disbursements))
            ws.cell(row, 7, float(variance))
            ws.cell(row, 8, float(variance_pct))
            ws.cell(row, 9, float(entry.obligation_rate))
            ws.cell(row, 10, float(entry.disbursement_rate))

            # Audit findings
            findings = self._generate_audit_findings(entry, variance, variance_pct)
            ws.cell(row, 11, findings)

            # Number formatting
            for col in [4, 5, 6, 7]:
                ws.cell(row, col).number_format = '#,##0.00'
            for col in [8, 9, 10]:
                ws.cell(row, col).number_format = '0.00'

            row += 1

        # Adjust column widths
        self._adjust_variance_column_widths(ws)

    def _write_activity_breakdown(self, ws, queryset):
        """Write activity-level budget breakdown sheet."""
        # Header
        ws.merge_cells('A1:I1')
        ws['A1'] = "ACTIVITY-LEVEL BUDGET BREAKDOWN"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Column headers
        row = 4
        headers = [
            "PPA Code",
            "PPA Title",
            "Activity/Task",
            "Activity Type",
            "Allocated Budget (₱)",
            "Actual Expenditure (₱)",
            "Variance (₱)",
            "Variance (%)",
            "Status"
        ]
        self._write_column_headers(ws, row, headers)

        # Data rows
        row += 1
        for entry in queryset:
            if not entry.enable_workitem_tracking or not entry.execution_project:
                # No WorkItem tracking, show PPA-level only
                allocated = entry.budget_allocation or Decimal('0.00')
                disbursements = entry.total_disbursements
                variance = allocated - disbursements
                variance_pct = (variance / allocated * 100) if allocated > 0 else 0

                ws.cell(row, 1, entry.program_code or f"PPA-{entry.id.hex[:8]}")
                ws.cell(row, 2, entry.title)
                ws.cell(row, 3, "PPA Level (No Activity Breakdown)")
                ws.cell(row, 4, "N/A")
                ws.cell(row, 5, float(allocated))
                ws.cell(row, 6, float(disbursements))
                ws.cell(row, 7, float(variance))
                ws.cell(row, 8, float(variance_pct))
                ws.cell(row, 9, entry.get_status_display())

                # Number formatting
                for col in [5, 6, 7]:
                    ws.cell(row, col).number_format = '#,##0.00'
                ws.cell(row, 8).number_format = '0.00'

                row += 1
            else:
                # Has WorkItem tracking, show activity breakdown
                activities = entry.execution_project.get_children()

                for activity in activities:
                    allocated = activity.allocated_budget or Decimal('0.00')
                    actual = activity.actual_expenditure or Decimal('0.00')
                    variance = allocated - actual
                    variance_pct = (variance / allocated * 100) if allocated > 0 else 0

                    ws.cell(row, 1, entry.program_code or f"PPA-{entry.id.hex[:8]}")
                    ws.cell(row, 2, entry.title)
                    ws.cell(row, 3, activity.title)
                    ws.cell(row, 4, activity.get_work_type_display())
                    ws.cell(row, 5, float(allocated))
                    ws.cell(row, 6, float(actual))
                    ws.cell(row, 7, float(variance))
                    ws.cell(row, 8, float(variance_pct))
                    ws.cell(row, 9, activity.get_status_display())

                    # Number formatting
                    for col in [5, 6, 7]:
                        ws.cell(row, col).number_format = '#,##0.00'
                    ws.cell(row, 8).number_format = '0.00'

                    row += 1

        # Adjust column widths
        column_widths = {
            'A': 15, 'B': 35, 'C': 40, 'D': 18, 'E': 18,
            'F': 20, 'G': 15, 'H': 12, 'I': 15
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    def _write_audit_trail(self, ws, queryset):
        """Write audit trail export sheet using django-auditlog."""
        # Header
        ws.merge_cells('A1:H1')
        ws['A1'] = "AUDIT TRAIL EXPORT"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Column headers
        row = 4
        headers = [
            "Timestamp",
            "PPA Code",
            "Action",
            "Field Changed",
            "Old Value",
            "New Value",
            "User",
            "IP Address"
        ]
        self._write_column_headers(ws, row, headers)

        # Get ContentType for MonitoringEntry
        content_type = ContentType.objects.get_for_model(MonitoringEntry)

        # Query audit logs for these entries
        entry_ids = list(queryset.values_list('id', flat=True))
        audit_logs = LogEntry.objects.filter(
            content_type=content_type,
            object_id__in=[str(eid) for eid in entry_ids]
        ).select_related('actor').order_by('-timestamp')[:1000]  # Limit to 1000 recent entries

        # Data rows
        row += 1
        for log in audit_logs:
            # Get PPA code from changes
            ppa_code = "N/A"
            try:
                entry = MonitoringEntry.objects.get(id=log.object_id)
                ppa_code = entry.program_code or f"PPA-{entry.id.hex[:8]}"
            except MonitoringEntry.DoesNotExist:
                pass

            # Parse changes
            changes = log.changes_dict if hasattr(log, 'changes_dict') else {}

            if changes:
                for field, (old_val, new_val) in changes.items():
                    ws.cell(row, 1, log.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
                    ws.cell(row, 2, ppa_code)
                    ws.cell(row, 3, log.get_action_display())
                    ws.cell(row, 4, field)
                    ws.cell(row, 5, str(old_val)[:50])  # Truncate long values
                    ws.cell(row, 6, str(new_val)[:50])
                    ws.cell(row, 7, log.actor.get_full_name() if log.actor else "System")
                    ws.cell(row, 8, log.remote_addr or "N/A")
                    row += 1
            else:
                # No detailed changes, just log the action
                ws.cell(row, 1, log.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
                ws.cell(row, 2, ppa_code)
                ws.cell(row, 3, log.get_action_display())
                ws.cell(row, 4, "N/A")
                ws.cell(row, 5, "N/A")
                ws.cell(row, 6, "N/A")
                ws.cell(row, 7, log.actor.get_full_name() if log.actor else "System")
                ws.cell(row, 8, log.remote_addr or "N/A")
                row += 1

        # Adjust column widths
        column_widths = {
            'A': 20, 'B': 15, 'C': 15, 'D': 20, 'E': 25,
            'F': 25, 'G': 20, 'H': 15
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    def _write_column_headers(self, ws, row, headers):
        """Write column headers with formatting."""
        header_fill = PatternFill(start_color='8B0000', end_color='8B0000', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

    def _generate_audit_findings(self, entry, variance, variance_pct) -> str:
        """Generate audit findings based on variance analysis."""
        findings = []

        if abs(variance_pct) > 20:
            findings.append("MATERIAL VARIANCE")

        if entry.is_over_budget:
            findings.append("BUDGET OVERRUN")

        if entry.is_under_obligated:
            findings.append("LOW OBLIGATION RATE")

        if entry.obligation_rate > 100:
            findings.append("OVER-OBLIGATION")

        if entry.approval_status != 'approved':
            findings.append(f"PENDING {entry.get_approval_status_display().upper()}")

        # Check for funding flow discrepancies
        allocation_variance = entry.allocation_variance
        if abs(allocation_variance) > Decimal('1000.00'):  # Significant variance
            findings.append("ALLOCATION DISCREPANCY")

        return "; ".join(findings) if findings else "NO FINDINGS"

    def _adjust_variance_column_widths(self, ws):
        """Auto-adjust column widths for variance summary."""
        column_widths = {
            'A': 15, 'B': 40, 'C': 30, 'D': 18, 'E': 18,
            'F': 20, 'G': 15, 'H': 12, 'I': 15, 'J': 15, 'K': 30
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width


# Export functions for easy import
def generate_mfbm_budget_execution_report(fiscal_year: int, moa_filter: Optional[str] = None) -> io.BytesIO:
    """
    Generate MFBM Budget Execution Report.

    Args:
        fiscal_year: Fiscal year for the report
        moa_filter: Optional MOA organization ID to filter by

    Returns:
        BytesIO: Excel file buffer
    """
    generator = MFBMBudgetExecutionReportGenerator(fiscal_year, moa_filter)
    return generator.generate()


def generate_bpda_development_report(fiscal_year: int, sector: Optional[str] = None) -> io.BytesIO:
    """
    Generate BPDA Development Alignment Report.

    Args:
        fiscal_year: Fiscal year for the report
        sector: Optional sector filter

    Returns:
        BytesIO: Excel file buffer
    """
    generator = BPDADevelopmentReportGenerator(fiscal_year, sector)
    return generator.generate()


def generate_coa_variance_report(fiscal_year: int, moa_filter: Optional[str] = None) -> io.BytesIO:
    """
    Generate COA Budget Variance Report.

    Args:
        fiscal_year: Fiscal year for the report
        moa_filter: Optional MOA organization ID to filter

    Returns:
        BytesIO: Excel file buffer with variance analysis and audit trail
    """
    generator = COAVarianceReportGenerator(fiscal_year, moa_filter)
    return generator.generate()
