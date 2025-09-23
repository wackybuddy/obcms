# Data Import/Export/Report Functions for OBC Communities

import csv
import io
import json
from datetime import datetime

import pandas as pd
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.views.decorators.http import require_http_methods
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table,
                                TableStyle)

from .models import OBCCommunity


@login_required
@require_http_methods(["POST"])
def import_communities_csv(request):
    """Import communities from CSV/Excel file."""
    try:
        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            return JsonResponse({"error": "No file uploaded"}, status=400)

        # Check file extension
        file_extension = uploaded_file.name.split(".")[-1].lower()
        if file_extension not in ["csv", "xlsx", "xls"]:
            return JsonResponse(
                {"error": "Unsupported file format. Please upload CSV or Excel files."},
                status=400,
            )

        # Read the file
        try:
            if file_extension == "csv":
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)
        except Exception as e:
            return JsonResponse({"error": f"Error reading file: {str(e)}"}, status=400)

        # Validate required columns
        required_columns = ["barangay_id", "community_names"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return JsonResponse(
                {
                    "error": f'Missing required columns: {", ".join(missing_columns)}',
                    "available_columns": list(df.columns),
                },
                status=400,
            )

        # Process and save communities
        created_count = 0
        updated_count = 0
        errors = []

        for index, row in df.iterrows():
            try:
                # Get barangay
                from common.models import Barangay

                barangay = Barangay.objects.get(id=row["barangay_id"])

                # Create or update community
                community_data = {
                    "barangay": barangay,
                    "community_names": row.get("community_names", ""),
                    "obc_id": row.get("obc_id", ""),
                    "estimated_obc_population": (
                        row.get("estimated_obc_population")
                        if pd.notna(row.get("estimated_obc_population"))
                        else None
                    ),
                    "households": (
                        row.get("households")
                        if pd.notna(row.get("households"))
                        else None
                    ),
                    "primary_ethnolinguistic_group": row.get(
                        "primary_ethnolinguistic_group", ""
                    ),
                    "unemployment_rate": row.get("unemployment_rate", "unknown"),
                    "settlement_type": row.get("settlement_type", "village"),
                    "latitude": (
                        row.get("latitude") if pd.notna(row.get("latitude")) else None
                    ),
                    "longitude": (
                        row.get("longitude") if pd.notna(row.get("longitude")) else None
                    ),
                }

                # Check if community exists
                existing = OBCCommunity.objects.filter(
                    barangay=barangay, community_names=row.get("community_names", "")
                ).first()

                if existing:
                    # Update existing community
                    for key, value in community_data.items():
                        if key != "barangay" and value is not None and value != "":
                            setattr(existing, key, value)
                    existing.save()
                    updated_count += 1
                else:
                    # Create new community
                    OBCCommunity.objects.create(**community_data)
                    created_count += 1

            except Exception as e:
                errors.append(f"Row {index + 1}: {str(e)}")

        result = {
            "success": True,
            "created": created_count,
            "updated": updated_count,
            "total_processed": len(df),
            "errors": errors[:10],  # Limit to first 10 errors
        }

        if errors:
            result["message"] = (
                f"Processed {created_count + updated_count} communities with {len(errors)} errors"
            )
        else:
            result["message"] = (
                f"Successfully processed {created_count + updated_count} communities"
            )

        return JsonResponse(result)

    except Exception as e:
        return JsonResponse({"error": f"Unexpected error: {str(e)}"}, status=500)


@login_required
def export_communities(request):
    """Export communities to Excel/CSV/PDF."""
    export_format = request.GET.get("format", "excel")
    include_fields = request.GET.getlist("fields[]") or "all"

    # Get communities queryset
    communities = OBCCommunity.objects.select_related(
        "barangay",
        "barangay__municipality",
        "barangay__municipality__province",
        "barangay__municipality__province__region",
    ).all()

    try:
        if export_format == "excel":
            return _export_to_excel(communities, include_fields)
        elif export_format == "csv":
            return _export_to_csv(communities, include_fields)
        elif export_format == "pdf":
            return _export_to_pdf(communities, include_fields)
        else:
            return JsonResponse({"error": "Invalid export format"}, status=400)
    except Exception as e:
        return JsonResponse({"error": f"Export error: {str(e)}"}, status=500)


def _export_to_excel(communities, include_fields):
    """Export to Excel format."""
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "OBC Communities"

    # Define field mappings
    field_mappings = {
        "basic": [
            ("OBC ID", "obc_id"),
            ("Community Names", "community_names"),
            ("Region", "barangay__municipality__province__region__name"),
            ("Province", "barangay__municipality__province__name"),
            ("Municipality", "barangay__municipality__name"),
            ("Barangay", "barangay__name"),
            ("Settlement Type", "get_settlement_type_display"),
            ("Population", "estimated_obc_population"),
            ("Households", "households"),
        ],
        "demographics": [
            ("Children (0-9)", "children_0_9"),
            ("Adolescents (10-14)", "adolescents_10_14"),
            ("Youth (15-30)", "youth_15_30"),
            ("Adults (31-59)", "adults_31_59"),
            ("Seniors (60+)", "seniors_60_plus"),
            ("Primary Ethno Group", "get_primary_ethnolinguistic_group_display"),
            ("Languages", "languages_spoken"),
        ],
        "services": [
            ("Access Education", "get_access_formal_education_display"),
            ("Access Healthcare", "get_access_healthcare_display"),
            ("Access Water", "get_access_clean_water_display"),
            ("Access Electricity", "get_access_electricity_display"),
        ],
        "location": [
            ("Latitude", "latitude"),
            ("Longitude", "longitude"),
            ("Proximity BARMM", "get_proximity_to_barmm_display"),
        ],
    }

    # Determine which fields to include
    if include_fields == "all" or "all" in include_fields:
        selected_fields = []
        for category in field_mappings.values():
            selected_fields.extend(category)
    else:
        selected_fields = []
        for field_category in include_fields:
            if field_category in field_mappings:
                selected_fields.extend(field_mappings[field_category])

    # Create headers
    headers = [field[0] for field in selected_fields]
    ws.append(headers)

    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Add data rows
    for community in communities:
        row_data = []
        for field in selected_fields:
            value = _get_field_value(community, field[1])
            row_data.append(value)
        ws.append(row_data)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="obc_communities_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    )

    # Save workbook to response
    wb.save(response)
    return response


def _export_to_csv(communities, include_fields):
    """Export to CSV format."""
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = (
        f'attachment; filename="obc_communities_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    )

    writer = csv.writer(response)

    # Define basic fields for CSV (simplified)
    headers = [
        "OBC ID",
        "Community Names",
        "Region",
        "Province",
        "Municipality",
        "Barangay",
        "Settlement Type",
        "Population",
        "Households",
        "Primary Ethno Group",
        "Latitude",
        "Longitude",
        "Development Status",
    ]

    writer.writerow(headers)

    for community in communities:
        row = [
            community.obc_id or "",
            community.community_names or "",
            community.barangay.municipality.province.region.name,
            community.barangay.municipality.province.name,
            community.barangay.municipality.name,
            community.barangay.name,
            community.get_settlement_type_display(),
            community.estimated_obc_population or "",
            community.households or "",
            community.get_primary_ethnolinguistic_group_display() or "",
            community.latitude or "",
            community.longitude or "",
            community.get_unemployment_rate_display(),
        ]
        writer.writerow(row)

    return response


def _export_to_pdf(communities, include_fields):
    """Export to PDF format."""
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="obc_communities_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    )

    # Create PDF document
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=16,
        spaceAfter=30,
        alignment=1,  # Center alignment
    )
    title = Paragraph("OBC Communities Report", title_style)
    elements.append(title)

    # Summary statistics
    total_communities = communities.count()
    total_population = sum(c.estimated_obc_population or 0 for c in communities)
    total_households = sum(c.households or 0 for c in communities)

    summary_data = [
        ["Total Communities", str(total_communities)],
        ["Total Population", f"{total_population:,}"],
        ["Total Households", f"{total_households:,}"],
        ["Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
    ]

    summary_table = Table(summary_data, colWidths=[2 * inch, 2 * inch])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )

    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # Communities table (simplified for PDF)
    table_data = [
        ["Community", "Location", "Population", "Households", "Development Status"]
    ]

    for community in communities[:50]:  # Limit to first 50 for PDF
        location = f"{community.barangay.name}, {community.barangay.municipality.name}"
        table_data.append(
            [
                (
                    community.community_names[:30] + "..."
                    if len(community.community_names or "") > 30
                    else community.community_names or "N/A"
                ),
                location[:35] + "..." if len(location) > 35 else location,
                str(community.estimated_obc_population or "N/A"),
                str(community.households or "N/A"),
                community.get_unemployment_rate_display(),
            ]
        )

    data_table = Table(
        table_data, colWidths=[1.5 * inch, 2 * inch, 1 * inch, 1 * inch, 1.5 * inch]
    )
    data_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("FONTSIZE", (0, 1), (-1, -1), 8),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )

    elements.append(data_table)

    if communities.count() > 50:
        note = Paragraph(
            f"<i>Note: Showing first 50 of {communities.count()} communities. Use Excel export for complete data.</i>",
            styles["Normal"],
        )
        elements.append(Spacer(1, 20))
        elements.append(note)

    doc.build(elements)
    return response


@login_required
def generate_obc_report(request):
    """Generate comprehensive OBC data report."""
    report_type = request.GET.get("type", "summary")
    format_type = request.GET.get("format", "pdf")

    try:
        if report_type == "summary":
            return _generate_summary_report(format_type)
        elif report_type == "demographic":
            return _generate_demographic_report(format_type)
        elif report_type == "geographic":
            return _generate_geographic_report(format_type)
        else:
            return JsonResponse({"error": "Invalid report type"}, status=400)
    except Exception as e:
        return JsonResponse({"error": f"Report generation error: {str(e)}"}, status=500)


def _generate_summary_report(format_type):
    """Generate summary report of OBC communities."""
    communities = OBCCommunity.objects.select_related(
        "barangay__municipality__province__region"
    ).all()

    if format_type == "pdf":
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="obc_summary_report_{datetime.now().strftime("%Y%m%d")}.pdf"'
        )

        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title = Paragraph("OBC Communities Summary Report", styles["Title"])
        elements.append(title)
        elements.append(Spacer(1, 20))

        # Executive Summary
        exec_summary = f"""
        <b>Executive Summary</b><br/><br/>
        This report provides a comprehensive overview of Other Bangsamoro Communities (OBC) 
        located outside the Bangsamoro Autonomous Region in Muslim Mindanao (BARMM).<br/><br/>
        <b>Key Statistics:</b><br/>
        • Total Communities: {communities.count()}<br/>
        • Total Population: {sum(c.estimated_obc_population or 0 for c in communities):,}<br/>
        • Total Households: {sum(c.households or 0 for c in communities):,}<br/>
        • Coverage: Regions IX and XII primarily<br/><br/>
        """

        elements.append(Paragraph(exec_summary, styles["Normal"]))
        elements.append(Spacer(1, 20))

        # Regional distribution
        from django.db.models import Count

        regional_data = (
            communities.values("barangay__municipality__province__region__name")
            .annotate(count=Count("id"), total_pop=Count("estimated_obc_population"))
            .order_by("-count")
        )

        region_table_data = [["Region", "Communities", "Est. Population"]]
        for item in regional_data:
            region_name = item["barangay__municipality__province__region__name"]
            community_count = item["count"]
            # Calculate actual population sum for this region
            region_pop = sum(
                c.estimated_obc_population or 0
                for c in communities.filter(
                    barangay__municipality__province__region__name=region_name
                )
            )
            region_table_data.append(
                [region_name, str(community_count), f"{region_pop:,}"]
            )

        region_table = Table(
            region_table_data, colWidths=[3 * inch, 1.5 * inch, 1.5 * inch]
        )
        region_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        elements.append(Paragraph("<b>Regional Distribution</b>", styles["Heading2"]))
        elements.append(region_table)
        elements.append(Spacer(1, 20))

        # Unemployment rate distribution
        unemployment_data = {}
        for community in communities:
            rate = community.get_unemployment_rate_display()
            unemployment_data[rate] = unemployment_data.get(rate, 0) + 1

        status_table_data = [["Unemployment Rate", "Count", "Percentage"]]
        total = communities.count()
        for rate, count in unemployment_data.items():
            percentage = (count / total * 100) if total > 0 else 0
            status_table_data.append([rate, str(count), f"{percentage:.1f}%"])

        status_table = Table(
            status_table_data, colWidths=[2.5 * inch, 1 * inch, 1 * inch]
        )
        status_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, 0), 12),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                    ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
                    ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ]
            )
        )

        elements.append(
            Paragraph("<b>Development Status Distribution</b>", styles["Heading2"])
        )
        elements.append(status_table)

        doc.build(elements)
        return response

    else:  # Excel format
        return _export_summary_to_excel(communities)


def _export_summary_to_excel(communities):
    """Export summary report to Excel."""
    wb = Workbook()

    # Summary sheet
    ws1 = wb.active
    ws1.title = "Summary"

    # Basic stats
    ws1["A1"] = "OBC Communities Summary Report"
    ws1["A1"].font = Font(size=16, bold=True)

    ws1["A3"] = "Total Communities:"
    ws1["B3"] = communities.count()
    ws1["A4"] = "Total Population:"
    ws1["B4"] = sum(c.estimated_obc_population or 0 for c in communities)
    ws1["A5"] = "Total Households:"
    ws1["B5"] = sum(c.households or 0 for c in communities)

    # Regional breakdown
    ws1["A7"] = "Regional Distribution:"
    ws1["A7"].font = Font(bold=True)

    row = 8
    regional_stats = {}
    for community in communities:
        region = community.barangay.municipality.province.region.name
        if region not in regional_stats:
            regional_stats[region] = {"count": 0, "population": 0}
        regional_stats[region]["count"] += 1
        regional_stats[region]["population"] += community.estimated_obc_population or 0

    for region, stats in regional_stats.items():
        ws1[f"A{row}"] = region
        ws1[f"B{row}"] = stats["count"]
        ws1[f"C{row}"] = stats["population"]
        row += 1

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="obc_summary_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    )

    wb.save(response)
    return response


@login_required
def data_guidelines(request):
    """Display data collection guidelines."""
    return render(request, "common/data_guidelines.html")


def _get_field_value(obj, field_path):
    """Get field value from object using dot notation or method calls."""
    try:
        if "__" in field_path:
            # Handle related field access
            parts = field_path.split("__")
            value = obj
            for part in parts:
                if hasattr(value, part):
                    value = getattr(value, part)
                else:
                    return "N/A"
            return value if value is not None else "N/A"
        elif field_path.startswith("get_") and field_path.endswith("_display"):
            # Handle choice field display methods
            if hasattr(obj, field_path):
                return getattr(obj, field_path)()
            return "N/A"
        else:
            # Handle direct field access
            value = getattr(obj, field_path, None)
            return value if value is not None else "N/A"
    except Exception:
        return "N/A"


def _generate_demographic_report(format_type):
    """Generate demographic analysis report."""
    # Implementation for demographic report
    pass


def _generate_geographic_report(format_type):
    """Generate geographic distribution report."""
    # Implementation for geographic report
    pass
